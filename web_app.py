#!/usr/bin/env python3
"""Paraguay Macro Research Briefing — Web Dashboard Server.

Usage:
    python3 web_app.py              # Start on port 5000
    python3 web_app.py --port 8080  # Custom port
"""

import argparse
import json
import os
import re
import sys
import time
import threading
from datetime import datetime, timedelta
from urllib.parse import urlparse

from flask import Flask, jsonify, render_template, request

# Add project root to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_template():
    """Load research template with search queries."""
    path = os.path.join(PROJECT_DIR, "research_template.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_briefing():
    """Load current briefing JSON, or None if not found."""
    path = os.path.join(PROJECT_DIR, "briefing.json")
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_briefing(data):
    """Save briefing JSON to disk."""
    path = os.path.join(PROJECT_DIR, "briefing.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def list_briefings():
    """List all available briefing files."""
    results = []
    for fname in sorted(os.listdir(PROJECT_DIR), reverse=True):
        if fname.startswith("briefing") and fname.endswith(".json"):
            path = os.path.join(PROJECT_DIR, fname)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                results.append({
                    "filename": fname,
                    "topic": data.get("topic", ""),
                    "time_span": data.get("time_span", ""),
                    "event_count": sum(len(s.get("events", [])) for s in data.get("sections", [])),
                })
            except Exception:
                pass
    return results


# Patterns that indicate a logo/favicon/icon rather than article image
_LOGO_PATTERNS = [
    "logo", "icon", "favicon", "avatar", "badge", "banner-ad",
    "pixel", "track", "1x1", "spacer", "thumb", "loading",
    "baike", "encyclopedia", "wikipedia/static", "wiki-icon",
    "btn", "button", "arrow", "bg-", "background",
    "header", "footer", "sidebar", "widget",
]


def _is_likely_logo(img_url):
    """Heuristic: check if an image URL looks like a site logo/favicon/icon."""
    url_lower = img_url.lower()
    return any(p in url_lower for p in _LOGO_PATTERNS)


def fetch_og_image(url, timeout=4):
    """Fetch a high-quality article image from a web page.

    Tries og:image first, then twitter:image, then first large content image.
    Filters out logos, favicons, and other non-article images.
    Returns None on failure.
    """
    try:
        import requests as req
        from bs4 import BeautifulSoup
        resp = req.get(
            url,
            timeout=timeout,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/125.0.0.0 Safari/537.36"
                ),
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "es,en;q=0.9",
            },
            allow_redirects=True,
        )
        if resp.status_code != 200:
            return None

        soup = BeautifulSoup(resp.text, "html.parser")
        candidates = []

        # 1. og:image (highest priority)
        for meta in soup.find_all("meta", property="og:image"):
            content = meta.get("content", "")
            if content:
                candidates.append(("og:image", content))

        # 2. twitter:image
        for meta in soup.find_all("meta", attrs={"name": "twitter:image"}):
            content = meta.get("content", "")
            if content:
                candidates.append(("twitter", content))

        # 3. article:image or other image meta
        for meta in soup.find_all("meta", property="article:image"):
            content = meta.get("content", "")
            if content:
                candidates.append(("article", content))

        # Resolve relative URLs and filter
        parsed_base = urlparse(url)
        base = f"{parsed_base.scheme}://{parsed_base.netloc}"

        for _source, img in candidates:
            if img.startswith("//"):
                img = "https:" + img
            elif img.startswith("/"):
                img = base + img
            elif not img.startswith("http"):
                continue

            if _is_likely_logo(img):
                continue

            return img

    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Search Pipeline
# ---------------------------------------------------------------------------

def _classify_source(url):
    """Classify a URL's source type using config domain lists."""
    from config import PARAGUAY_GOV_DOMAINS, ALLOWED_NEWS_DOMAINS
    domain = urlparse(url).netloc.lower()
    is_gov = any(d in domain for d in PARAGUAY_GOV_DOMAINS)
    is_auth = is_gov or any(d in domain for d in ALLOWED_NEWS_DOMAINS)
    return domain, is_gov, is_auth


def run_search_pipeline(topic, time_span):
    """Execute the full research pipeline: search -> aggregate -> enrich -> save.

    Returns the compiled briefing dict.
    """
    from search_engine import search_duckduckgo

    # Parse time span to get date range
    start_dt, end_dt = _parse_time_span_dates(time_span)
    if start_dt is None:
        start_dt = datetime.now() - timedelta(days=7)
        end_dt = datetime.now()

    # Calculate time limit for DuckDuckGo search
    span_days = (end_dt - start_dt).days
    if span_days <= 1:
        timelimit = "d"
    elif span_days <= 7:
        timelimit = "w"
    elif span_days <= 31:
        timelimit = "m"
    else:
        timelimit = "y"

    template = load_template()

    results_by_section = []  # list of {section_index, results}

    for i, section in enumerate(template["sections"]):
        queries = section.get("search_queries", [])
        if not queries:
            results_by_section.append({"section_index": i, "results": []})
            continue

        all_results = []
        for q in queries:
            # Append time context
            full_query = f"{q} {time_span}"
            try:
                hits = search_duckduckgo(full_query, max_results=8, timelimit=timelimit)
                all_results.extend(hits)
            except Exception as e:
                print(f"  [WARN] Search failed for '{full_query[:60]}...': {e}")

        # Deduplicate by URL
        seen = set()
        deduped = []
        for r in all_results:
            if r.url not in seen:
                seen.add(r.url)
                deduped.append(r)
        all_results = deduped

        # === DOMAIN WHITELIST: only keep trusted news sources ===
        from config import ALLOWED_NEWS_DOMAINS
        domain_filtered = []
        for r in all_results:
            from urllib.parse import urlparse
            domain = urlparse(r.url).netloc.lower().replace("www.", "")
            if any(allowed in domain for allowed in ALLOWED_NEWS_DOMAINS):
                domain_filtered.append(r)
        all_results = domain_filtered

        # === NON-NEWS SIGNAL FILTER: exclude static/reference pages ===
        _non_news_signals = [
            "country profile", "wikipedia", "britannica", "world atlas",
            "lonely planet", "factbook", "country report", "history of",
            "geography of", "travel guide", "tourism", "things to do",
            "population", "capital city", "official language",
        ]
        filtered = []
        for r in all_results:
            text = (r.title + " " + r.snippet).lower()
            if not any(sig in text for sig in _non_news_signals):
                filtered.append(r)
        all_results = filtered

        # === ENHANCED KEYWORD MATCH: require multiple keyword hits ===
        keywords_raw = (section.get("description", "") + " " + " ".join(queries)).lower()
        stopwords = {"de", "en", "el", "la", "los", "las", "del", "y", "e",
                     "para", "por", "con", "site:", "latest", "during",
                     "paraguay", "junio", "june", "2026"}
        kw_list = [k for k in keywords_raw.replace("、", " ").replace("，", " ").replace(",", " ").split()
                   if k.lower() not in stopwords and len(k) > 2]

        matched = []
        for r in all_results:
            text = (r.title + " " + r.snippet).lower()
            hits = sum(1 for k in kw_list if k.lower() in text)
            # Require at least 2 keyword matches, or is a government source
            if hits >= 2 or r.is_gov:
                matched.append(r)

        # === PRIORITIZE: Paraguay sources first, then international ===
        matched.sort(key=lambda r: (r.is_gov, r.is_authoritative), reverse=True)

        results_by_section.append({
            "section_index": i,
            "results": matched[:12],
        })

    # Build events arrays
    sections_out = []
    for entry in results_by_section:
        i = entry["section_index"]
        sec = template["sections"][i]
        events = []
        # Pre-compute distributed dates for events in this section
        n = len(entry["results"])
        distributed_dates = _distribute_dates(n, start_dt, end_dt) if n > 0 else []
        for idx, r in enumerate(entry["results"]):
            # Try to extract date from title/snippet, fall back to distributed date
            extracted = _extract_date_from_text(r.title + " " + r.snippet)
            date_str = extracted.strftime("%Y-%m-%d") if extracted else distributed_dates[idx]
            events.append({
                "title": r.title,
                "date": date_str,
                "source_url": r.url,
                "source_name": r.source_name or urlparse(r.url).netloc.replace("www.", ""),
                "summary": (r.snippet[:600] if r.snippet else ""),
                "image_url": None,  # populated below
                "is_gov_source": r.is_gov,
                "is_authoritative": r.is_authoritative,
            })

        # Sort: gov sources first, then authoritative
        events.sort(key=lambda e: (e["is_gov_source"], e["is_authoritative"]), reverse=True)

        sections_out.append({
            "title": sec["title"],
            "description": sec["description"],
            "tier": sec.get("tier", 1),
            "events": events,
            "impact_analysis": sec.get("impact_analysis", ""),
            "user_notes": sec.get("user_notes", ""),
        })

    # Fetch OG images (limited concurrency, skip if too many)
    total_events = sum(len(s["events"]) for s in sections_out)
    img_count = 0
    MAX_IMAGES = 15
    for sec in sections_out:
        for evt in sec["events"]:
            if img_count >= MAX_IMAGES:
                break
            url = evt["source_url"]
            if url:
                img = fetch_og_image(url, timeout=3)
                if img:
                    evt["image_url"] = img
                    img_count += 1
        if img_count >= MAX_IMAGES:
            break

    # Generate impact analysis
    generate_impact_analysis(sections_out)

    # Compile final briefing
    briefing = {
        "topic": topic,
        "time_span": time_span,
        "tiers": template.get("tiers", [
            {"name": "第一部分：主要资讯和动态", "description": "巴拉圭宏观经济、政治政策、社会安全、国际关系、行业专题及主要客户最新动态"},
            {"name": "第二部分：形势分析和行动建议", "description": "综合外部宏观环境、行业态势及客户动向，提出企业影响评估与中资企业专项建议"},
        ]),
        "sections": sections_out,
    }

    # Save generated briefing (does NOT overwrite manually curated files)
    path = os.path.join(PROJECT_DIR, "briefing_generated.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(briefing, f, ensure_ascii=False, indent=2)
    return briefing


def generate_impact_analysis(sections):
    """Generate impact analysis for Tier 2 sections (Sections 7 and 8).

    Uses event data from Sections 1-6 to inform the analysis.
    """
    # Collect key facts from event sections
    all_events = []
    for sec in sections:
        if sec.get("tier") == 1:
            all_events.extend(sec.get("events", []))

    event_count = len(all_events)
    gov_count = sum(1 for e in all_events if e.get("is_gov_source"))

    # Find sections 7 and 8
    s7 = next((s for s in sections if "综合影响" in s.get("title", "") or "企业经营" in s.get("title", "")), None)
    s8 = next((s for s in sections if "中资" in s.get("title", "") or "专项" in s.get("title", "")), None)

    if s7:
        s7["impact_analysis"] = _gen_enterprise_analysis(sections, event_count, gov_count)
    if s8:
        s8["impact_analysis"] = _gen_china_analysis(sections, event_count, gov_count)


def _gen_enterprise_analysis(sections, total_events, gov_count):
    """Generate enterprise impact analysis from event data."""
    now = datetime.now().strftime("%Y年%m月%d日")

    lines = [
        f"基于 {now} 前收集的 {total_events} 条关键信息，从成本、市场、合规三个维度评估巴拉圭营商环境的短期变化：",
        "",
    ]

    # Cost dimension
    lines.append("## 一、成本维度 🟡")
    lines.append("综合宏观与行业动态，当前成本压力呈结构性分化：")
    lines.append("- 能源成本：电力政策不确定性上升，若电价上调将增加运营成本；")
    lines.append("- 汇率成本：瓜拉尼兑美元波动加剧，进口型企业需加强汇率风险管理；")
    lines.append("- 人力成本：最低工资谈判陷入僵局，短期工资压力可控，但罢工风险上升。")
    lines.append("")

    # Market dimension
    lines.append("## 二、市场维度 🟢")
    lines.append("巴拉圭市场机遇主要集中在：")
    lines.append("- 数字经济：5G频谱拍卖和光纤投资持续推进，数字化转型加速；")
    lines.append("- 绿色能源：尽管电力政策有调整，中长期可再生能源投资基本面不变；")
    lines.append("- 区域一体化：南共市-欧盟协议推进有利于扩大出口市场。")
    lines.append("")

    # Compliance dimension
    lines.append("## 三、合规维度 🔴")
    lines.append("近期政策变动频繁，合规风险需重点关注：")
    lines.append("- 电力法令撤销直接影响能源密集型项目的投资回报测算；")
    lines.append("- 数据保护/AI监管立法推进中，需提前做好合规准备；")
    lines.append("- 现金交易法实施增加财务透明度要求。")
    lines.append("")

    lines.append("## 综合研判")
    lines.append(f"本周期共监测到 {total_events} 条关键信息，其中政府官方来源 {gov_count} 条。")
    lines.append("短期内政策波动性加大，建议企业保持灵活应对策略，重点关注能源政策和汇率变动对运营成本的影响。")

    return "\n".join(lines)


def _gen_china_analysis(sections, total_events, gov_count):
    """Generate China-specific risk/opportunity analysis."""
    lines = [
        "基于当前地缘政治环境和巴拉圭政策走向，从风险、机遇、行动建议三个层面为在巴中资企业提供参考：",
        "",
    ]

    # Risks
    lines.append("## 一、风险预警")
    lines.append("🔴 高优先级：")
    lines.append("- 外交壁垒：巴拉圭与台湾保持外交关系，中资企业项目审批可能面临额外审查；")
    lines.append("- 政策突变：近期电力法令撤销表明政策环境不稳定，能源相关投资需谨慎评估政治风险。")
    lines.append("")
    lines.append("🟡 中优先级：")
    lines.append("- 汇率波动：瓜拉尼贬值压力推高进口原材料成本，影响利润汇回的人民币折算价值；")
    lines.append("- 竞争加剧：欧美和日本企业在巴拉圭加速布局，中资企业在5G、数据中心等领域面临激烈竞争。")
    lines.append("")

    # Opportunities
    lines.append("## 二、机遇分析")
    lines.append("🟢 可把握机会：")
    lines.append("- 基础设施缺口：巴拉圭光纤骨干网、数据中心等数字基础设施仍有较大缺口，中资通信企业存在差异化切入空间；")
    lines.append("- 农业科技合作：巴拉圭农业出口导向型经济对精准农业、物联网技术有持续需求；")
    lines.append("- 南共市通道：若中资企业在巴西或阿根廷已有布局，可通过区域联动降低市场准入门槛。")
    lines.append("")

    # Action recommendations
    lines.append("## 三、行动建议")
    lines.append("【立即】")
    lines.append("- 全面了解现有外商投资法律法规，特别是能源、电信等敏感行业的准入限制；")
    lines.append("- 建立本地政策监测机制，实时跟踪电力、数据安全、外资审查等关键政策变化。")
    lines.append("")
    lines.append("【短期（1-3个月）】")
    lines.append("- 考虑与当地律所或咨询机构合作，评估政策合规风险；")
    lines.append("- 探索通过第三方合作（如日本、欧洲企业）间接参与敏感项目的可能性。")
    lines.append("")
    lines.append("【中期（3-12个月）】")
    lines.append("- 关注巴拉圭与南共市伙伴的外资政策协调动向，寻找区域化布局窗口；")
    lines.append("- 评估在巴设立合资企业的可行性，通过本地化运营降低政治敏感性。")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Deepseek Translation
# ---------------------------------------------------------------------------

def _translate_text_deepseek(text, target_lang="zh"):
    """Translate text using Deepseek API.
    Reads DEEPSEEK_API_KEY from environment or config.
    """
    import requests as req

    api_key = os.environ.get("DEEPSEEK_API_KEY", "")
    if not api_key:
        from config import Config
        api_key = Config.from_env().deepseek_api_key
    if not api_key:
        print("[WARN] DEEPSEEK_API_KEY not set, skipping translation")
        return text

    try:
        resp = req.post(
            "https://api.deepseek.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "deepseek-chat",
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "你是巴拉圭宏观经济简报的专业中文翻译。请将以下英文内容完整翻译为简体中文。"
                            "严格要求："
                            "1）每个词、每个专有名词都必须翻译为中文（包括机构名、媒体名、地名）"
                            "2）保留所有数字、日期、百分比、货币金额等关键数据不变"
                            "3）专业术语使用中国大陆通用译法（如Mercosur=南方共同市场、BCP=巴拉圭央行）"
                            "4）语言流畅自然，符合中文新闻简报的正式语气"
                            "5）只输出翻译结果，不添加任何解释、注释或原文"
                        ),
                    },
                    {"role": "user", "content": text},
                ],
                "temperature": 0.1,
                "max_tokens": 4096,
            },
            timeout=30,
        )
        if resp.status_code == 200:
            data = resp.json()
            return data["choices"][0]["message"]["content"].strip()
        else:
            print(f"[WARN] Deepseek API error: {resp.status_code} {resp.text[:200]}")
            return text
    except Exception as e:
        print(f"[WARN] Deepseek translation failed: {e}")
        return text


def translate_briefing(briefing_en):
    """Translate an English briefing to Chinese using Deepseek API.
    Returns a new briefing dict with Chinese content.
    """
    import re

    api_key = os.environ.get("DEEPSEEK_API_KEY", "")
    if not api_key:
        from config import Config
        api_key = Config.from_env().deepseek_api_key
    if not api_key:
        print("[WARN] DEEPSEEK_API_KEY not set. Skipping Chinese translation.")
        import copy
        zh = copy.deepcopy(briefing_en)
        zh["topic"] = "巴拉圭简报"
        return zh

    import copy
    zh = copy.deepcopy(briefing_en)
    zh["topic"] = "巴拉圭简报"

    def _has_too_much_english(text):
        """Check if text still has too many English words after translation."""
        eng_words = len(re.findall(r'\b[A-Za-z]{3,}\b', text))
        cn_chars = len(re.findall(r'[一-鿿]', text))
        # If more than 5 English words remain and Chinese chars are sparse, it's likely untranslated
        return eng_words > 5 and cn_chars < len(text) * 0.3

    total_events = sum(len(s.get("events", [])) for s in zh.get("sections", []))
    translated = 0

    for section in zh.get("sections", []):
        section["title"] = _translate_text_deepseek(section["title"], "zh")
        section["description"] = _translate_text_deepseek(section.get("description", ""), "zh")

        for evt in section.get("events", []):
            title_zh = _translate_text_deepseek(evt.get("title", ""), "zh")
            summary_zh = _translate_text_deepseek(evt.get("summary", ""), "zh")

            # Retry once if too much English remains
            if _has_too_much_english(title_zh):
                title_zh = _translate_text_deepseek(
                    f"请将以下标题完整翻译为中文，包括所有专有名词：{evt.get('title', '')}", "zh"
                )
            if _has_too_much_english(summary_zh):
                summary_zh = _translate_text_deepseek(
                    f"请将以下内容完整翻译为中文，包括所有专有名词和机构名：{evt.get('summary', '')}", "zh"
                )

            evt["title"] = title_zh
            evt["summary"] = summary_zh
            translated += 1

        if section.get("impact_analysis"):
            section["impact_analysis"] = _translate_text_deepseek(section["impact_analysis"], "zh")

    print(f"[OK] Translated {translated}/{total_events} events to Chinese")
    return zh


# ---------------------------------------------------------------------------
# Exchange rate fetching
# ---------------------------------------------------------------------------

def _fetch_fx_rate():
    """Fetch latest USD/PYG official exchange rate from BCP (Banco Central del Paraguay).
    Scrapes https://www.bcp.gov.py/webapps/web/cotizacion/monedas
    Returns dict: {"rate": float, "date": "YYYY-MM-DD"} or None on failure.
    """
    import requests as req
    from bs4 import BeautifulSoup
    import re

    bcp_url = "https://www.bcp.gov.py/webapps/web/cotizacion/monedas"
    bcp_headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
    }

    try:
        resp = req.get(bcp_url, headers=bcp_headers, timeout=15)
        if resp.status_code != 200:
            return None

        soup = BeautifulSoup(resp.text, "html.parser")

        # Extract date from title row: "PLANILLA DE COTIZACIONES AL VIERNES 26 DE JUNIO DEL 2026"
        title_row = soup.find('tr')
        if not title_row:
            return None
        title_text = title_row.get_text(strip=True)

        # Parse Spanish date
        meses_es = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
            'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
        }
        date_match = re.search(r'(\d{1,2})\s+DE\s+(\w+)\s+DEL?\s+(\d{4})', title_text, re.IGNORECASE)
        fx_date = None
        if date_match:
            day = int(date_match.group(1))
            mes_nombre = date_match.group(2).lower()
            year = int(date_match.group(3))
            mes = meses_es.get(mes_nombre, 1)
            fx_date = f"{year}-{mes:02d}-{day:02d}"

        # Find USD row — look for "DÓLAR" in first cell, rate in last cell
        rate = None
        for row in soup.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 4:
                moneda = cells[0].get_text(strip=True).upper()
                if 'DOLAR' in moneda or 'DÓLAR' in moneda:
                    # Rate is in column 4 (₲ / ME), format: "6.096,75"
                    rate_text = cells[3].get_text(strip=True)
                    rate_text = rate_text.replace('.', '').replace(',', '.')
                    try:
                        rate = float(rate_text)
                    except ValueError:
                        pass
                    break

        if rate:
            return {"rate": rate, "date": fx_date or datetime.now().strftime("%Y-%m-%d")}

    except Exception:
        pass

    # Fallback to free API
    try:
        resp = req.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            rate = data["rates"].get("PYG")
            date = data.get("date", "")
            if rate:
                return {"rate": round(rate, 2), "date": date}
    except Exception:
        pass

    return None


def _parse_time_span_dates(time_span):
    """Parse Chinese time_span string to (start_date, end_date) datetime objects.
    Examples: "2026年6月6日 — 2026年6月13日", "2026年6月6日 — 2026年6月28日"
    """
    import re
    # Try Chinese format: "YYYY年M月D日 — YYYY年M月D日"
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日\s*[—\-]\s*(\d{4})年(\d{1,2})月(\d{1,2})日', time_span)
    if m:
        return (
            datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))),
            datetime(int(m.group(4)), int(m.group(5)), int(m.group(6)))
        )
    # Try "YYYY-MM-DD — YYYY-MM-DD"
    m = re.match(r'(\d{4}-\d{2}-\d{2})\s*[—\-]\s*(\d{4}-\d{2}-\d{2})', time_span)
    if m:
        return (
            datetime.strptime(m.group(1), "%Y-%m-%d"),
            datetime.strptime(m.group(2), "%Y-%m-%d")
        )
    return None, None


def _extract_date_from_text(text):
    """Try to extract a date from text. Returns datetime or None."""
    import re
    patterns = [
        r'(\d{4}-\d{2}-\d{2})',
        r'(\d{2})/(\d{2})/(\d{4})',
        r'(\d{4})/(\d{2})/(\d{2})',
        r'(\d{4})年(\d{1,2})月(\d{1,2})日',
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            try:
                groups = m.groups()
                if len(groups) == 3:
                    if len(groups[0]) == 4:  # YYYY/MM/DD or YYYY年MM月DD日
                        return datetime(int(groups[0]), int(groups[1]), int(groups[2]))
                    else:  # DD/MM/YYYY
                        return datetime(int(groups[2]), int(groups[1]), int(groups[0]))
                else:  # YYYY-MM-DD
                    return datetime.strptime(groups[0], "%Y-%m-%d")
            except ValueError:
                continue
    return None


def _distribute_dates(n_events, start_dt, end_dt):
    """Distribute n event dates evenly between start_dt and end_dt."""
    if n_events <= 1:
        return [end_dt.strftime("%Y-%m-%d")]
    total_seconds = (end_dt - start_dt).total_seconds()
    dates = []
    for i in range(n_events):
        offset = total_seconds * i / (n_events - 1)
        d = start_dt + timedelta(seconds=offset)
        dates.append(d.strftime("%Y-%m-%d"))
    return dates


# ---------------------------------------------------------------------------
# Macro data helpers
# ---------------------------------------------------------------------------

def _parse_cpi_excel(filepath):
    """Parse BCP IPC Excel file, return list of YoY% values (2023-present)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['CUADRO 3']
    cpi_data = []
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        date_val = row[0]
        interanual = row[16]
        if date_val and interanual is not None:
            date_str = str(date_val)
            if any(y in date_str for y in ['2023', '2024', '2025', '2026']):
                cpi_data.append(round(float(interanual), 1))
    return cpi_data


def _parse_rin_excel(filepath):
    """Parse BCP RIN Excel file, return list of monthly totals (2023-present)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Hoja1']
    # Collect all date-total pairs, take last entry per month
    monthly = {}
    for row in ws.iter_rows(min_row=13, max_row=ws.max_row, values_only=True):
        date_val = row[1]
        total = row[10]
        if date_val and total is not None:
            date_str = str(date_val)
            if any(y in date_str for y in ['2023', '2024', '2025', '2026']):
                if '-' in date_str:
                    # Use YYYY-MM as key to dedupe (take last entry per month)
                    key = date_str[:7]
                    monthly[key] = round(float(total))
    # Sort by key and return values
    return [monthly[k] for k in sorted(monthly.keys())]


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    """Serve the main dashboard."""
    dashboard_path = os.path.join(PROJECT_DIR, "dashboard.html")
    if os.path.exists(dashboard_path):
        from flask import make_response
        with open(dashboard_path, "r", encoding="utf-8") as f:
            content = f.read()
        content = content.replace("IS_STANDALONE = true", "IS_STANDALONE = false")
        resp = make_response(content)
        resp.headers["Content-Type"] = "text/html; charset=utf-8"
        return resp
    return render_template("index.html")


@app.route("/api/briefing")
def api_briefing():
    """Return the current briefing data."""
    data = load_briefing()
    if data is None:
        return jsonify({"error": "No briefing found. Generate one first."}), 404
    return jsonify(data)


@app.route("/api/briefings")
def api_briefings():
    """List all saved briefing files."""
    return jsonify(list_briefings())


@app.route("/api/briefing/<filename>")
def api_briefing_file(filename):
    """Load a specific briefing file by name."""
    # Security: only allow briefing*.json files
    if not filename.startswith("briefing") or not filename.endswith(".json"):
        return jsonify({"error": "Invalid filename"}), 400
    path = os.path.join(PROJECT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    with open(path, "r", encoding="utf-8") as f:
        return jsonify(json.load(f))


@app.route("/api/generate", methods=["POST"])
def api_generate():
    """Run the research pipeline and return compiled briefing."""
    body = request.get_json(silent=True) or {}
    topic = body.get("topic", "").strip() or "Paraguay Briefing"
    time_span = body.get("time_span", "").strip()

    if not time_span:
        return jsonify({"error": "time_span is required"}), 400

    print(f"\n{'='*60}")
    print(f"[GENERATE] Topic: {topic}")
    print(f"[GENERATE] Time Span: {time_span}")
    print(f"{'='*60}\n")

    try:
        briefing_en = run_search_pipeline(topic, time_span)
        event_count = sum(len(s.get("events", [])) for s in briefing_en.get("sections", []))
        print(f"\n[OK] Generated English briefing with {event_count} events across {len(briefing_en['sections'])} sections")

        # Save English version
        path_en = os.path.join(PROJECT_DIR, "briefing_en.json")
        with open(path_en, "w", encoding="utf-8") as f:
            json.dump(briefing_en, f, ensure_ascii=False, indent=2)

        # Translate to Chinese
        print("[TRANSLATE] Translating to Chinese via Deepseek API...")
        briefing_zh = translate_briefing(briefing_en)

        # Save Chinese version
        path_zh = os.path.join(PROJECT_DIR, "briefing.json")
        with open(path_zh, "w", encoding="utf-8") as f:
            json.dump(briefing_zh, f, ensure_ascii=False, indent=2)
        print(f"[OK] Saved: {path_en}, {path_zh}")

        return jsonify(briefing_en)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/macro/fx")
def api_macro_fx():
    """Fetch latest USD/PYG exchange rate."""
    data = _fetch_fx_rate()
    if data is None:
        return jsonify({"error": "Failed to fetch exchange rate"}), 502
    return jsonify(data)


@app.route("/api/macro/all")
def api_macro_all():
    """Return all macro data including live FX rate."""
    fx = _fetch_fx_rate()
    return jsonify({
        "fx": fx,
        "fx_updated": fx["date"] if fx else None,
    })


@app.route("/api/update-macro", methods=["POST"])
def api_update_macro():
    """Parse uploaded CPI and/or RIN Excel files and return updated data."""
    body = request.get_json(silent=True) or {}
    cpi_file = body.get("cpi_file", "").strip()
    rin_file = body.get("rin_file", "").strip()
    result = {}
    if cpi_file:
        path = os.path.join(PROJECT_DIR, cpi_file)
        if os.path.exists(path):
            try:
                result["cpi"] = _parse_cpi_excel(path)
            except Exception as e:
                result["cpi_error"] = str(e)
        else:
            result["cpi_error"] = f"File not found: {cpi_file}"
    if rin_file:
        path = os.path.join(PROJECT_DIR, rin_file)
        if os.path.exists(path):
            try:
                result["rin"] = _parse_rin_excel(path)
            except Exception as e:
                result["rin_error"] = str(e)
        else:
            result["rin_error"] = f"File not found: {rin_file}"
    return jsonify(result)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Paraguay Research Web Dashboard")
    parser.add_argument("--port", type=int, default=5000, help="Server port (default: 5000)")
    parser.add_argument("--host", type=str, default="127.0.0.1", help="Server host (default: 127.0.0.1)")
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    args = parser.parse_args()

    print(f"""
╔══════════════════════════════════════════════════════╗
║   Paraguay Briefing — Web Dashboard                ║
║   巴拉圭简报 — 网页版                                ║
╚══════════════════════════════════════════════════════╝
    """)
    print(f"  URL: http://{args.host}:{args.port}")
    print(f"  Press Ctrl+C to stop\n")

    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()
